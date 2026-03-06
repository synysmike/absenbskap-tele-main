"""
Telegram bot for presensi.bskap.id: multi-user absen (teguh_in/teguh_out)
and scheduled absen with Excel + per-user image folders.
"""
import base64
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Optional

import telebot
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from apscheduler.schedulers.background import BackgroundScheduler
import requests

from test import PresensiClient

load_dotenv()

# --- Config ---
API_TOKEN = os.getenv("API_TOKEN")
PRESENSI_BASE_URL = "https://presensi.bskap.id"
KODE_JAM_KERJA = "JK04"

# User keys supported (lowercase)
USER_KEYS = ["teguh", "guntur", "ayu", "hisah", "widhi"]

def _load_user_credentials() -> Dict[str, Dict[str, str]]:
    """Load per-user presensi credentials from env. Keys: teguh, guntur, ..."""
    creds = {}
    for key in USER_KEYS:
        prefix = key.upper()
        user_id = os.getenv(f"{prefix}_USER_ID")
        password = os.getenv(f"{prefix}_PASSWORD")
        lokasi = os.getenv(f"{prefix}_LOKASI", "Jakarta")
        lokasi_cabang = os.getenv(f"{prefix}_LOKASI_CABANG", "-7.316514,112.724501")
        if user_id and password:
            creds[key] = {
                "user_id": user_id,
                "password": password,
                "lokasi": lokasi,
                "lokasi_cabang": lokasi_cabang,
            }
    return creds

USER_CREDENTIALS = _load_user_credentials()

# Commands: "teguh_in", "teguh_out", "teguh_in_now", "teguh_out_now", ...
IMMEDIATE_COMMANDS = []
for u in USER_KEYS:
    if u in USER_CREDENTIALS:
        IMMEDIATE_COMMANDS.append(f"{u}_in")
        IMMEDIATE_COMMANDS.append(f"{u}_out")
        IMMEDIATE_COMMANDS.append(f"{u}_in_now")
        IMMEDIATE_COMMANDS.append(f"{u}_out_now")

# Schedule Excel
SCHEDULES_EXCEL = Path("schedules.xlsx")
SHEET_NAME = "Schedules"
IMAGES_DIR = Path("images")

# Headers for schedule sheet
SCHEDULE_HEADERS = ["username", "status", "date", "time", "image_path", "done", "chat_id"]


def ensure_schedule_excel() -> None:
    if not SCHEDULES_EXCEL.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(SCHEDULE_HEADERS)
        wb.save(SCHEDULES_EXCEL)


def add_schedule(username: str, status: str, date: str, time: str, image_path: str, chat_id: int) -> None:
    ensure_schedule_excel()
    wb = load_workbook(SCHEDULES_EXCEL)
    ws = wb[SHEET_NAME]
    ws.append([username, status, date, time, image_path, "N", str(chat_id)])
    wb.save(SCHEDULES_EXCEL)


def _excel_date_str(val: Any) -> str:
    """Normalize Excel date cell to DDmonYYYY string (openpyxl may return date/datetime)."""
    if val is None:
        return ""
    if hasattr(val, "strftime") and hasattr(val, "month"):
        return f"{val.day:02d}{datetime(2000, val.month, 1).strftime('%b').lower()}{val.year}"
    return str(val).strip()


def _excel_time_str(val: Any) -> str:
    """Normalize Excel time cell to HH:MM string (openpyxl may return time/datetime)."""
    if val is None:
        return ""
    if hasattr(val, "strftime") and hasattr(val, "hour"):
        return val.strftime("%H:%M")
    return str(val).strip()


def get_due_schedules() -> list:
    """Return list of rows (as dict) where done=N and scheduled datetime <= now."""
    if not SCHEDULES_EXCEL.exists():
        return []
    wb = load_workbook(SCHEDULES_EXCEL, read_only=False)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers = SCHEDULE_HEADERS
    due = []
    now = datetime.now()
    for idx, row in enumerate(rows, start=2):
        if not row or len(row) < 7:
            continue
        d = dict(zip(headers, row))
        if str(d.get("done") or "").strip().upper() == "Y":
            continue
        date_str = _excel_date_str(d.get("date"))
        time_str = _excel_time_str(d.get("time"))
        if not date_str or not time_str:
            continue
        try:
            dt = parse_schedule_datetime(date_str, time_str)
            if dt <= now:
                due.append({"row": row, "data": d, "row_idx": idx})
        except Exception:
            continue
    return due


def parse_schedule_datetime(date_str: str, time_str: str) -> datetime:
    """Parse 09mar2026 and 07:17 into datetime."""
    months = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
              "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}
    m = re.match(r"(\d{1,2})([a-z]{3})(\d{4})", date_str.strip().lower())
    if not m:
        raise ValueError(f"Invalid date format: {date_str}")
    day, mon, year = int(m.group(1)), m.group(2), int(m.group(3))
    month = months.get(mon)
    if not month:
        raise ValueError(f"Invalid month: {mon}")
    t = re.match(r"(\d{1,2}):(\d{2})", time_str.strip())
    if not t:
        raise ValueError(f"Invalid time format: {time_str}")
    hour, minute = int(t.group(1)), int(t.group(2))
    return datetime(year, month, day, hour, minute)


def mark_schedule_done(row_idx: int) -> None:
    wb = load_workbook(SCHEDULES_EXCEL, read_only=False)
    ws = wb[SHEET_NAME]
    ws.cell(row=row_idx, column=6, value="Y")
    wb.save(SCHEDULES_EXCEL)


def save_schedule_image(username: str, date_str: str, time_str: str, image_bytes: bytes) -> Path:
    """Save image to images/<username>/<date>_<time>.png. Time without colon."""
    safe_time = time_str.replace(":", "")
    filename = f"{date_str}_{safe_time}.png"
    folder = IMAGES_DIR / username.lower()
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / filename
    with open(path, "wb") as f:
        f.write(image_bytes)
    return path


def photo_to_data_uri(image_bytes: bytes, mime: str = "image/png") -> str:
    b64 = base64.b64encode(image_bytes).decode("utf-8")
    return f"data:{mime};base64,{b64}"


# --- Bot ---
bot = telebot.TeleBot(API_TOKEN)
user_states: Dict[int, Dict[str, Any]] = {}

# Help text for /help command (Telegram Markdown)
def get_help_text() -> str:
    names = ", ".join(USER_CREDENTIALS.keys()) if USER_CREDENTIALS else "teguh, guntur, ..."
    return (
        "*Bot Presensi BSKAP*\n\n"
        "*Perintah:*\n"
        "/start — Panduan singkat\n"
        "/help — Bantuan ini\n\n"
        "*1. Absen langsung:*\n"
        "Kirim `teguh_in` / `teguh_in_now` (masuk) atau `teguh_out` / `teguh_out_now` (keluar), lalu foto selfie.\n"
        f"User: {names}\n\n"
        "*2. Jadwal absen:*\n"
        "Format: `nama_status;tanggal;jam`\n"
        "Contoh: `teguh_in;09mar2026;07:17`\n"
        "Tanggal: DDmonYYYY, Jam: HH:MM. Lalu kirim foto selfie. Absen otomatis di waktu jadwal.\n\n"
        "_Lokasi diacak 2–6 m dari cabang._"
    )


@bot.message_handler(commands=["start"])
def cmd_start(message: telebot.types.Message):
    names = ", ".join(USER_CREDENTIALS.keys()) if USER_CREDENTIALS else "teguh, guntur, ..."
    bot.reply_to(
        message,
        f"Bot Presensi BSKAP.\n\n"
        f"**Direct absen (sekarang):** kirim `teguh_in_now` atau `teguh_out_now` (atau `teguh_in`/`teguh_out`), lalu kirim foto selfie. User: {names}.\n\n"
        f"**Jadwal absen:** kirim format\n"
        f"`teguh_in;09mar2026;07:17`\n"
        f"(nama_status;tanggal;jam). Lalu kirim foto selfie. Absen akan otomatis di waktu yang dijadwalkan.",
        parse_mode="Markdown",
    )


@bot.message_handler(commands=["help"])
def cmd_help(message: telebot.types.Message):
    bot.reply_to(message, get_help_text(), parse_mode="Markdown")


def is_immediate_command(text: str) -> bool:
    return text and text.strip().lower() in [c.lower() for c in IMMEDIATE_COMMANDS]


def parse_immediate_command(text: str) -> Optional[tuple[str, str]]:
    """From 'teguh_in' or 'teguh_in_now' return (username, status) or None."""
    t = (text or "").strip().lower()
    if t not in [c.lower() for c in IMMEDIATE_COMMANDS]:
        return None
    t = t.replace("_now", "")
    if "_" not in t:
        return None
    username, status = t.rsplit("_", 1)
    if username in USER_CREDENTIALS and status in ("in", "out"):
        return (username, status)
    return None


def parse_schedule_command(text: str) -> Optional[Dict[str, str]]:
    """Parse 'teguh_in;09mar2026;07:17' -> {username, status, date, time} or None."""
    text = (text or "").strip()
    parts = text.split(";")
    if len(parts) != 3:
        return None
    name_status, date_str, time_str = parts[0].strip(), parts[1].strip(), parts[2].strip()
    if "_" not in name_status:
        return None
    username, status = name_status.rsplit("_", 1)
    username = username.lower()
    status = status.lower()
    if username not in USER_CREDENTIALS or status not in ("in", "out"):
        return None
    try:
        parse_schedule_datetime(date_str, time_str)
    except Exception:
        return None
    return {"username": username, "status": status, "date": date_str, "time": time_str}


@bot.message_handler(func=lambda m: is_immediate_command(m.text))
def handle_immediate_absen(message: telebot.types.Message):
    parsed = parse_immediate_command(message.text)
    if not parsed:
        bot.reply_to(message, "User tidak terdaftar.")
        return
    username, status = parsed
    user_states[message.chat.id] = {
        "action": "immediate",
        "username": username,
        "status": status,
    }
    bot.reply_to(message, "Kirim foto selfie untuk absen sekarang (direct).")


@bot.message_handler(func=lambda m: parse_schedule_command(m.text))
def handle_schedule_absen(message: telebot.types.Message):
    parsed = parse_schedule_command(message.text)
    if not parsed:
        return
    user_states[message.chat.id] = {
        "action": "schedule",
        "username": parsed["username"],
        "status": parsed["status"],
        "date": parsed["date"],
        "time": parsed["time"],
    }
    bot.reply_to(
        message,
        f"Jadwal: {parsed['username']}_{parsed['status']} pada {parsed['date']} {parsed['time']}. Kirim foto selfie untuk menyimpan jadwal.",
    )


@bot.message_handler(content_types=["photo"])
def handle_photo(message: telebot.types.Message):
    chat_id = message.chat.id
    if chat_id not in user_states:
        bot.reply_to(message, "Pilih dulu: absen sekarang (teguh_in/teguh_out) atau buat jadwal (teguh_in;09mar2026;07:17).")
        return

    state = user_states[chat_id]
    file_info = bot.get_file(message.photo[-1].file_id)
    downloaded = bot.download_file(file_info.file_path)
    image_bytes = downloaded

    if state["action"] == "schedule":
        username = state["username"]
        date_str, time_str = state["date"], state["time"]
        try:
            path = save_schedule_image(username, date_str, time_str, image_bytes)
            add_schedule(username, state["status"], date_str, time_str, str(path), chat_id)
            bot.reply_to(
                message,
                f"Jadwal tersimpan. Absen {username}_{state['status']} akan dilakukan pada {date_str} {time_str}. Foto disimpan di {path}.",
            )
        except Exception as e:
            bot.reply_to(message, f"Gagal menyimpan jadwal: {e}")
        del user_states[chat_id]
        return

    # immediate absen
    username = state["username"]
    status = state["status"]
    creds = USER_CREDENTIALS.get(username)
    if not creds:
        bot.reply_to(message, "User tidak terdaftar.")
        del user_states[chat_id]
        return

    bot.reply_to(message, "Memproses absen...")
    try:
        client = PresensiClient(
            user_id=creds["user_id"],
            password=creds["password"],
            base_url=PRESENSI_BASE_URL,
        )
        client.login()
        data_uri = photo_to_data_uri(image_bytes)
        result = client.submit_presensi(
            image_data_uri=data_uri,
            status=status,
            lokasi=creds["lokasi"],
            lokasi_cabang=creds["lokasi_cabang"],
            kode_jam_kerja=KODE_JAM_KERJA,
        )
        if result.get("status") is True:
            bot.reply_to(message, f"Absen berhasil. {result.get('message', '')}")
        else:
            bot.reply_to(message, f"Absen gagal: {result.get('message', repr(result))}")
    except requests.exceptions.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else "unknown"
        body = ""
        try:
            body = e.response.text[:500] if e.response is not None else ""
        except Exception:
            body = ""
        bot.reply_to(
            message,
            f"HTTP error saat absen (status {status_code}). "
            f"Body: {body or 'tanpa body / tidak dapat dibaca.'}",
        )
    except Exception as e:
        bot.reply_to(message, f"Error: {e}")
    finally:
        del user_states[chat_id]


def run_scheduled_absens() -> None:
    """Check Excel for due schedules, run presensi, mark done, notify chat."""
    due = get_due_schedules()
    for item in due:
        d = item["data"]
        row_idx = item["row_idx"]
        image_path = (d.get("image_path") or "").strip()
        chat_id_str = (d.get("chat_id") or "").strip()
        username = (d.get("username") or "").strip().lower()
        status = (d.get("status") or "").strip().lower()
        if not image_path or not os.path.isfile(image_path) or username not in USER_CREDENTIALS:
            mark_schedule_done(row_idx)
            if chat_id_str:
                try:
                    bot.send_message(int(chat_id_str), f"Jadwal absen gagal: file tidak ditemukan atau user tidak valid.")
                except Exception:
                    pass
            continue
        creds = USER_CREDENTIALS[username]
        try:
            with open(image_path, "rb") as f:
                image_bytes = f.read()
            data_uri = photo_to_data_uri(image_bytes)
            client = PresensiClient(
                user_id=creds["user_id"],
                password=creds["password"],
                base_url=PRESENSI_BASE_URL,
            )
            client.login()
            result = client.submit_presensi(
                image_data_uri=data_uri,
                status=status,
                lokasi=creds["lokasi"],
                lokasi_cabang=creds["lokasi_cabang"],
                kode_jam_kerja=KODE_JAM_KERJA,
            )
            mark_schedule_done(row_idx)
            msg = f"Jadwal absen {username}_{status} selesai. {result.get('message', '')}" if result.get("status") else f"Jadwal absen gagal: {result.get('message', repr(result))}"
            if chat_id_str:
                try:
                    bot.send_message(int(chat_id_str), msg)
                except Exception:
                    pass
        except requests.exceptions.HTTPError as e:
            mark_schedule_done(row_idx)
            status_code = e.response.status_code if e.response is not None else "unknown"
            body = ""
            try:
                body = e.response.text[:500] if e.response is not None else ""
            except Exception:
                body = ""
            if chat_id_str:
                try:
                    bot.send_message(
                        int(chat_id_str),
                        f"Jadwal absen HTTP error (status {status_code}). "
                        f"Body: {body or 'tanpa body / tidak dapat dibaca.'}",
                    )
                except Exception:
                    pass
        except Exception as e:
            mark_schedule_done(row_idx)
            if chat_id_str:
                try:
                    bot.send_message(int(chat_id_str), f"Jadwal absen error: {e}")
                except Exception:
                    pass


# --- Scheduler ---
scheduler = BackgroundScheduler()
scheduler.add_job(run_scheduled_absens, "interval", minutes=1, id="presensi_schedule")
scheduler.start()

if __name__ == "__main__":
    if not API_TOKEN:
        print("Set API_TOKEN in .env (copy from .env.example and fill).")
        raise SystemExit(1)
    ensure_schedule_excel()
    print("Bot presensi running. Scheduler: every 1 min.")
    bot.infinity_polling()
